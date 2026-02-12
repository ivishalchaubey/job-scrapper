from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiParameter

from . import services
from .serializers import (
    JobSerializer, JobListSerializer, ScrapingRunSerializer,
    CompanyStatsSerializer, DashboardStatsSerializer,
)


@extend_schema(
    responses=JobListSerializer(many=True),
    parameters=[
        OpenApiParameter('company_name', str, description='Filter by company name'),
        OpenApiParameter('city', str, description='Filter by city'),
        OpenApiParameter('country', str, description='Filter by country'),
        OpenApiParameter('employment_type', str, description='Filter by employment type'),
        OpenApiParameter('department', str, description='Filter by department'),
        OpenApiParameter('search', str, description='Search across title, company, city, department'),
        OpenApiParameter('ordering', str, description='Sort field (prefix with - for desc)'),
        OpenApiParameter('page', int, description='Page number (default: 1)'),
        OpenApiParameter('page_size', int, description='Results per page (default: 50)'),
    ],
    description="List all active jobs with filtering, search, and pagination"
)
@api_view(['GET'])
def job_list_view(request):
    filters = {}
    for key in ['company_name', 'city', 'country', 'employment_type', 'department']:
        val = request.query_params.get(key)
        if val:
            filters[key] = val

    search = request.query_params.get('search', '')
    ordering = request.query_params.get('ordering', '-updated_at')
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 50))

    jobs, total = services.get_jobs(
        filters=filters, search=search, ordering=ordering,
        page=page, page_size=page_size,
    )

    return Response({
        'count': total,
        'page': page,
        'page_size': page_size,
        'results': jobs,
    })


@extend_schema(responses=JobSerializer, description="Get a single job by ID")
@api_view(['GET'])
def job_detail_view(request, job_id):
    job = services.get_job_by_id(job_id)
    if not job:
        return Response({'error': 'Job not found'}, status=404)
    return Response(job)


@extend_schema(
    responses=DashboardStatsSerializer,
    description="Get dashboard statistics including total jobs, active companies, and success rate"
)
@api_view(['GET'])
def stats_view(request):
    return Response(services.get_dashboard_stats())


@extend_schema(
    responses=CompanyStatsSerializer(many=True),
    description="Get all companies with their job counts and last scrape date"
)
@api_view(['GET'])
def companies_view(request):
    return Response(services.get_company_stats())


@extend_schema(
    responses=ScrapingRunSerializer(many=True),
    parameters=[OpenApiParameter('limit', int, description='Number of records (default: 50)')],
    description="Get recent scraping run history"
)
@api_view(['GET'])
def scraping_history_view(request):
    limit = int(request.query_params.get('limit', 50))
    return Response(services.get_scraping_history(limit=limit))


@extend_schema(description="Health check endpoint")
@api_view(['GET'])
def health_view(request):
    from core.db import get_db
    try:
        get_db().command('ping')
        db_status = 'connected'
    except Exception:
        db_status = 'disconnected'

    return Response({
        'status': 'healthy' if db_status == 'connected' else 'degraded',
        'database': db_status,
        'scrapers': 275,
        'timestamp': datetime.now().isoformat(),
    })


@extend_schema(description="Delete jobs by list of IDs")
@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def delete_jobs_view(request):
    job_ids = request.data.get('job_ids', [])
    if not job_ids:
        return Response({'error': 'No job IDs provided'}, status=400)
    deleted = services.delete_jobs_by_ids(job_ids)
    return Response({'deleted': deleted})


@extend_schema(description="Clear all jobs and scraping runs from database")
@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def clear_all_view(request):
    services.delete_all_jobs()
    return Response({'status': 'ok', 'message': 'All data cleared'})


@extend_schema(
    parameters=[
        OpenApiParameter('company_name', str, description='Filter by company name'),
        OpenApiParameter('city', str, description='Filter by city'),
        OpenApiParameter('country', str, description='Filter by country'),
        OpenApiParameter('search', str, description='Search across title, company, city, department'),
    ],
    description="Export jobs as XLSX file"
)
@api_view(['GET'])
def export_xlsx_view(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filters = {}
    for key in ['company_name', 'city', 'country', 'employment_type', 'department']:
        val = request.query_params.get(key)
        if val:
            filters[key] = val

    search = request.query_params.get('search', '')
    jobs, total = services.get_jobs(
        filters=filters, search=search, ordering='-updated_at',
        page=1, page_size=50000,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Jobs'

    headers = [
        'Company', 'Title', 'Location', 'City', 'Country',
        'Department', 'Employment Type', 'Experience Level',
        'Salary Range', 'Posted Date', 'Apply URL',
    ]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, job in enumerate(jobs, 2):
        row_data = [
            job.get('company_name', ''),
            job.get('title', ''),
            job.get('location', ''),
            job.get('city', ''),
            job.get('country', ''),
            job.get('department', ''),
            job.get('employment_type', ''),
            job.get('experience_level', ''),
            job.get('salary_range', ''),
            job.get('posted_date', ''),
            job.get('apply_url', ''),
        ]
        for col_idx, value in enumerate(row_data, 1):
            if isinstance(value, (list, tuple)):
                value = ', '.join(str(v).strip() for v in value)
            elif not isinstance(value, str):
                value = str(value) if value is not None else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    col_widths = [20, 45, 30, 15, 12, 20, 15, 15, 15, 15, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'jobs_export_{timestamp}.xlsx'

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
